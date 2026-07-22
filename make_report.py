# /// script
# requires-python = ">=3.10"
# dependencies = ["boto3", "openpyxl", "requests"]
# ///
"""Export creative-tester job results to an Excel file matching the ad-quality report layout.

Usage:
    uv run make_report.py job_id [job_id ...] [-o report.xlsx]

Needs AWS credentials in the environment (same ones you use for sigcurl).
Reads the DynamoDB table directly (the API's /results route can't return large jobs —
277 rows of raw_result exceeds Lambda's 6 MB response cap).
DSP-only columns (Creative ID, Leased/Served Ads, Exchange, CPM, Buyer, ...) are left
blank — this service only has the classifier's output per URL.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from urllib.parse import urlparse, parse_qs, unquote
from zoneinfo import ZoneInfo

import boto3
from boto3.dynamodb.conditions import Key
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

REGION = "us-east-1"
TABLE = "creative-tester-dev-jobs"
CST = ZoneInfo("America/Chicago")
CELL_MAX = 32000  # Excel hard cap is 32767 chars per cell

COLUMNS = [
    "Creative ID", "Ad Url", "MIME Type", "Duration (Secs)", "Leased Ads", "Served Ads",
    "Dopped Ads", "Exchange", "Leased Timestamp (CST)", "First Approval Timestamp (CST)",
    "First Rejection Timestamp (CST)", "State Change Reason", "State", "Advertiser Domains",
    "IAB Categories", "Buyer", "Ad Rating", "Campaign Names", "Ad Title", "Ad Systems",
    "bitrate", "Orientation", "admJbCmpltTs", "Analysis of the Ad",
    "Text Extracted from Frames", "Tags", "Political/Social Sensitivity Score (0-10)",
    "Ad Analysis Status", "CPM", "Ad Tech Partners", "Currency", "Buyer Seat",
    # extra columns from this service, for filtering
    "Job ID", "Endpoint", "Row Status", "Completed (CST)", "Raw Result JSON",
]


def fetch_results(table, job_id: str) -> list[dict]:
    items, kwargs = [], {
        "KeyConditionExpression": Key("pk").eq(f"JOB#{job_id}") & Key("sk").begins_with("RESULT#"),
    }
    while True:
        resp = table.query(**kwargs)
        items.extend(resp.get("Items", []))
        last = resp.get("LastEvaluatedKey")
        if not last:
            return items
        kwargs["ExclusiveStartKey"] = last


def first(raw: dict, *keys, default=""):
    for k in keys:
        v = raw.get(k)
        if v not in (None, ""):
            return v
    return default


def ad_title(url: str) -> str:
    # the real media URL may be tucked inside a tracker redirect
    q = parse_qs(urlparse(url).query)
    if "redirect_url" in q:
        url = unquote(q["redirect_url"][0])
    name = urlparse(url).path.rsplit("/", 1)[-1]
    name = name.rsplit(".", 1)[0]
    return name.replace("_", " ").replace("-", " ").strip()


def cst(epoch_s) -> str:
    if not epoch_s:
        return ""
    dt = datetime.fromtimestamp(int(epoch_s), tz=timezone.utc).astimezone(CST)
    return dt.strftime("%d %b %Y, %I:%M %p").lstrip("0")


def clip(v) -> str:
    s = str(v)
    return s[:CELL_MAX] + "…[truncated]" if len(s) > CELL_MAX else s


def row_for(job_id: str, item: dict) -> dict:
    raw = item.get("raw_result") or {}
    kind = item.get("asset_kind", "")
    url = item.get("url", "")
    mime = "video/mp4" if kind == "video" else ("image/*" if kind == "image" else "")
    completed = item.get("completed_at")
    tags = item.get("norm_tags") or []
    return {
        "Ad Url": url,
        "MIME Type": mime,
        "Duration (Secs)": first(raw, "duration", "duration_secs", "duration_seconds", "video_duration"),
        "Ad Title": ad_title(url),
        "bitrate": first(raw, "bitrate", "bit_rate"),
        "Orientation": first(raw, "orientation"),
        "admJbCmpltTs": int(completed) * 1000 if completed else "",
        "Analysis of the Ad": clip(first(raw, "analysis", "ad_analysis", "analysis_of_the_ad", "summary", "reasoning", "description")),
        "Text Extracted from Frames": clip(first(raw, "extracted_text", "text_extracted_from_frames", "ocr_text", "frame_text", "text")),
        "Tags": ", ".join(tags),
        "Political/Social Sensitivity Score (0-10)": item.get("score", first(raw, "political_social_ad_likelihood")),
        "Ad Analysis Status": str(first(raw, "status", default=item.get("status", ""))).upper(),
        "Job ID": job_id,
        "Endpoint": item.get("endpoint", ""),
        "Row Status": item.get("status", ""),
        "Completed (CST)": cst(completed),
        "Raw Result JSON": clip(json.dumps(raw, default=str)) if raw else "",
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("job_ids", nargs="+")
    ap.add_argument("-o", "--out", default="creative_report.xlsx")
    args = ap.parse_args()

    table = boto3.resource("dynamodb", region_name=REGION).Table(TABLE)

    rows, per_job = [], {}
    for jid in args.job_ids:
        items = fetch_results(table, jid)
        per_job[jid] = items
        rows.extend(row_for(jid, it) for it in items)
        done = sum(1 for it in items if it.get("status") == "completed")
        print(f"{jid}: {len(items)} rows ({done} completed)")

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(col, "") for col in COLUMNS])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows) + 1, 2)}"
    ws.freeze_panes = "C2"
    widths = {"Ad Url": 60, "Ad Title": 40, "Analysis of the Ad": 60,
              "Text Extracted from Frames": 60, "Tags": 25, "Raw Result JSON": 40}
    for i, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 14)

    # summary: tagged counts, overall and per tag / per job
    s = wb.create_sheet("Summary")
    s.append(["Job ID", "Rows", "Completed", "Tagged", "Untagged (completed)"])
    for c in s[1]:
        c.font = Font(bold=True)
    tag_counts: dict[str, int] = {}
    for jid, items in per_job.items():
        comp = [it for it in items if it.get("status") == "completed"]
        tagged = [it for it in comp if it.get("norm_tags")]
        for it in tagged:
            for t in it["norm_tags"]:
                tag_counts[t] = tag_counts.get(t, 0) + 1
        s.append([jid, len(items), len(comp), len(tagged), len(comp) - len(tagged)])
    s.append([])
    s.append(["Tag", "Count"])
    for c in s[s.max_row]:
        c.font = Font(bold=True)
    for t, n in sorted(tag_counts.items(), key=lambda kv: -kv[1]):
        s.append([t, n])
    for col, w in (("A", 34), ("B", 10), ("C", 12), ("D", 10), ("E", 20)):
        s.column_dimensions[col].width = w

    wb.save(args.out)
    total_tagged = sum(1 for r in rows if r["Tags"])
    print(f"\nWrote {args.out}: {len(rows)} rows, {total_tagged} tagged")
    for t, n in sorted(tag_counts.items(), key=lambda kv: -kv[1]):
        print(f"  {t}: {n}")


if __name__ == "__main__":
    sys.exit(main())
